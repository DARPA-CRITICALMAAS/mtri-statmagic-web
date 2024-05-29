import dm_util
import openpyxl

con1 = dm_util.con1

xlpath = './Cross-TA CMA synchronization document.xlsx'

# Mapping of "Evidence Layer" column names to database table fields
cmap = {
    'Dataset Name': 'name',
    'Alt Name': 'name_alt',
    'Description': 'description',
    'Source': 'source',
    'Link': 'path',

}

wb = openpyxl.load_workbook(xlpath)
sql = ''
unique_paths = []
for sheet in wb.worksheets:
    print(sheet.title)

    # Get start and end row indices
    index_start = 0
    index_end = 0
    for i,row in enumerate(sheet.rows):
        if row[0].value and 'Evidence Layer' in row[0].value:
            index_start = i+3

        if row[0].value and'Deposit Site' in row[0].value:
            index_end = i-2

    # Get hdr row
    hdr = [sheet.cell(index_start-1,j).value for j in range(1,sheet.max_column+1)]

    # If sheet is missing columns, skip
    skip = False
    for k, d in cmap.items():
        if k not in hdr:
            skip = True

    if skip:
        continue

    # print(
    #     index_start,
    #     index_end,
    #     sheet.cell(index_start,3).value,
    #     sheet.cell(index_end,2).value
    # )

    for i in range(index_start,index_end+1):
        #row = sheet.cell
        method = sheet.cell(i,hdr.index('Method')+1).value.split(' - ')
        data = {
            'category': method[0],
        }
        if len(method) > 1:
            data['subcategory'] = method[1]
        for k,d in cmap.items():
            v = sheet.cell(i,hdr.index(k)+1).value
            data[d] = v.replace("'",'"') if v else ''

        # Just load .tifs for now
        if '.tif' not in data['path']:
            continue

        data['data_format'] = 'tif'

        if data['path'] in unique_paths:
            continue

        unique_paths.append(data['path'])



        forder = sorted(list(data.keys()))
        cols = ','.join(forder)
        vals = "','".join([data[k] for k in forder])
        sql += f'''
            INSERT INTO datalayer ({cols})
            VALUES ('{vals}');
        '''
        #blerg

print(sql)
con1.query(sql)